import json
import os
from docx import Document
from openpyxl import load_workbook
from utils import get_openai_client, detect_language, translate_text

def process_docx(file_path: str, client) -> dict:
    """Process a .docx file and return its content with metadata."""
    doc = Document(file_path)
    content = "\n".join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])
    
    # Detect and translate content if not in English
    detected_lang = detect_language(content)
    if detected_lang != 'English':
        content = translate_text(client, content)
    
    return {
        "content": content,
        "metadata": {
            "source": "official_document",
            "file_name": os.path.basename(file_path),
            "file_type": "docx",
            "original_language": detected_lang
        }
    }

def process_xlsx(file_path: str, client) -> dict:
    """Process a .xlsx file and return its content with metadata."""
    wb = load_workbook(file_path)
    content = []
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Get column headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Start with sheet name
        content.append(f"\nSheet: {sheet}")
        
        # If this is a table-like structure with headers
        if headers:
            # Process each row
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = []
                for idx, cell in enumerate(row):
                    if cell is not None and str(cell).strip():
                        # Include header name for context
                        if idx < len(headers):
                            row_data.append(f"{headers[idx]}: {cell}")
                        else:
                            row_data.append(str(cell))
                
                if row_data:
                    # Join with newlines for better readability
                    content.append("\n".join(row_data))
        else:
            # For non-table sheets, try to preserve the structure
            for row in ws.iter_rows(values_only=True):
                row_content = []
                for cell in row:
                    if cell is not None and str(cell).strip():
                        row_content.append(str(cell))
                
                if row_content:
                    # If it's a single cell with a lot of text, preserve newlines
                    if len(row_content) == 1 and "\n" in row_content[0]:
                        content.append(row_content[0])
                    else:
                        content.append(" | ".join(row_content))
    
    full_content = "\n\n".join(content)
    
    # Detect and translate content if not in English
    detected_lang = detect_language(full_content)
    if detected_lang != 'English':
        full_content = translate_text(client, full_content)
    
    return {
        "content": full_content,
        "metadata": {
            "source": "official_document",
            "file_name": os.path.basename(file_path),
            "file_type": "xlsx",
            "original_language": detected_lang,
            "sheets": wb.sheetnames
        }
    }

def process_documents(input_files: list) -> list:
    """Process multiple document files and return their contents with metadata."""
    client = get_openai_client()
    processed_documents = []
    
    for file_path in input_files:
        try:
            if file_path.endswith('.docx'):
                processed = process_docx(file_path, client)
            elif file_path.endswith('.xlsx'):
                processed = process_xlsx(file_path, client)
            else:
                print(f"Unsupported file type: {file_path}")
                continue
                
            processed_documents.append(processed)
            print(f"Successfully processed: {file_path}")
            
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
    
    return processed_documents

def main():
    # List of documents to process
    input_files = [
        "Customer Succes Manual.docx",
        "FAQ supply NL 07-02-24.docx",
        "Reclaim Guideline 2025.xlsx",
        "Reclaim manual per Reason 2025.docx"
    ]
    
    # Process all documents
    processed_documents = process_documents(input_files)
    
    # Save processed documents to JSON
    with open("processed_documents.json", "w", encoding="utf-8") as f:
        json.dump(processed_documents, f, ensure_ascii=False, indent=2)
    
    print(f"\nProcessed {len(processed_documents)} documents and saved to processed_documents.json")

if __name__ == "__main__":
    main() 